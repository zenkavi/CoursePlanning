import json
import os
from io import BytesIO
from itertools import groupby

import openpyxl
import yaml
from flask import Flask, jsonify, redirect, render_template, request, send_file, url_for
from openpyxl.styles import Font, PatternFill

from config import load_config
from data_loader import load_courses, load_faculty
from load_calc import all_faculty_loads, count_key, new_preps_in_semester
from models import Assignment, Plan
from solver import solve as run_solver

app = Flask(__name__)

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
PLAN_FILE = os.path.join(DATA_DIR, "plan.json")
CONFIG_FILE = os.path.join(DATA_DIR, "config.yaml")

# Load static data once at startup
faculty_list = load_faculty()
courses = load_courses()
cfg = load_config()

_BOLD = Font(bold=True)
_STATUS_FILL = {
    "green":        PatternFill("solid", fgColor="C6EFCE"),
    "yellow-over":  PatternFill("solid", fgColor="FFEB9C"),
    "yellow-under": PatternFill("solid", fgColor="FFEB9C"),
    "red":          PatternFill("solid", fgColor="FFC7CE"),
}

COURSE_ORDER = [
    "sci10", "sci30a", "sci30b", "sci31a", "sci31b", "sci40", "sci50",
    "sci111", "sci112", "udl_lab_1", "udl_lab_2", "udl_lec_1", "udl_lec_2",
]


def load_plan() -> Plan:
    try:
        with open(PLAN_FILE) as f:
            return Plan.from_dict(json.load(f))
    except FileNotFoundError:
        return Plan()


def save_plan(plan: Plan):
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(PLAN_FILE, "w") as f:
        json.dump(plan.to_dict(), f, indent=2)


def compute_violations(faculty_list, loads, plan, courses, cfg) -> dict:
    """Return {faculty_name: {items, has_error, has_warning, count}} for active violations."""
    lab_cats = {"upper_div_lecture_lab", "upper_div_lab"}
    junior_cap = cfg.get("junior_faculty_hard_cap", 2.0)
    senior_cap = cfg.get("senior_faculty_soft_cap", 2.0)

    result = {}
    for faculty in faculty_list:
        fname = faculty.name
        items = []

        # ── Semester load cap violations ──────────────────────────────
        for (y, sem), data in loads.get(fname, {}).items():
            total = data.get("total", 0.0)
            label = f"{'Fall' if sem == 'fall' else 'Spring'} Y{y}"
            if faculty.is_junior() and total > junior_cap:
                items.append({
                    "type": "hard_cap", "severity": "error",
                    "description": f"{label}: load {total:.2f} exceeds junior hard cap ({junior_cap:.1f})",
                })
            elif faculty.is_senior() and total > senior_cap:
                items.append({
                    "type": "soft_cap", "severity": "warning",
                    "description": f"{label}: load {total:.2f} exceeds senior soft cap ({senior_cap:.1f})",
                })

        # ── Junior: > 1 new lab prep per academic year ────────────────
        if faculty.is_junior():
            cumulative = dict(faculty.prior_teaching_counts)
            for year in range(plan.year_range[0], plan.year_range[1] + 1):
                year_new_labs = set()
                for sem in ("fall", "spring"):
                    sem_assigns = [
                        a for a in plan.assignments
                        if a.faculty_name == fname and a.year == year and a.semester == sem
                    ]
                    for a in sem_assigns:
                        course = courses.get(a.course_code)
                        if course and course.category in lab_cats:
                            if cumulative.get(count_key(a), 0) == 0:
                                year_new_labs.add(a.course_code)
                    for a in sem_assigns:
                        ck = count_key(a)
                        cumulative[ck] = cumulative.get(ck, 0) + 1

                if len(year_new_labs) > 1:
                    names = ", ".join(
                        courses[c].display_name if c in courses else c
                        for c in sorted(year_new_labs)
                    )
                    items.append({
                        "type": "new_lab_prep", "severity": "warning",
                        "description": (
                            f"Y{year}: {len(year_new_labs)} new lab preps ({names})"
                            f" — max 1/year for junior faculty"
                        ),
                    })

        result[fname] = {
            "items": items,
            "has_error": any(v["severity"] == "error" for v in items),
            "has_warning": any(v["severity"] == "warning" for v in items),
            "count": len(items),
        }
    return result


def _sci10_count(plan: Plan, year: int, season: str) -> int:
    key = f"{year}__{season}"
    default = courses["sci10"].sections_in(season) if "sci10" in courses else 10
    return plan.sci10_section_overrides.get(key, default)


def build_grid(plan: Plan, year_range: tuple = (1, 3)) -> list:
    """Return the semester/slot structure the template renders.

    Each semester dict:
      label, year, season, groups: [{course, sections: [{slot_id, section_number, assignment}]}]
    """
    semesters = []
    for year in range(year_range[0], year_range[1] + 1):
        for season in ("fall", "spring"):
            groups = []
            for code in COURSE_ORDER:
                course = courses.get(code)
                if course is None:
                    continue
                n = _sci10_count(plan, year, season) if code == "sci10" else course.sections_in(season)
                if n == 0:
                    continue
                sections = []
                for sec in range(1, n + 1):
                    sections.append({
                        "slot_id": f"{code}__{year}__{season}__{sec}",
                        "section_number": sec,
                        "assignment": plan.get_assignment(code, year, season, sec),
                    })
                groups.append({"course": course, "sections": sections})
            semesters.append({
                "year": year,
                "season": season,
                "label": f"{'Fall' if season == 'fall' else 'Spring'} Y{year}",
                "groups": groups,
                "total_sections": sum(len(g["sections"]) for g in groups),
            })
    return semesters


@app.route("/")
def index():
    plan = load_plan()
    semesters = build_grid(plan)
    loads = all_faculty_loads(faculty_list, plan.assignments, courses, cfg)
    # Flatten loads for Jinja: "year__season" -> {faculty_name: {total, status, ...}}
    loads_by_sem = {}
    for fname, sems in loads.items():
        for (y, sem), data in sems.items():
            loads_by_sem.setdefault(f"{y}__{sem}", {})[fname] = data

    # Annual totals with status: {faculty_name: {year: {total, status}}}
    annual_loads = {}
    faculty_by_name_map = {f.name: f for f in faculty_list}
    for fname, sems in loads.items():
        f = faculty_by_name_map.get(fname)
        annual_loads[fname] = {}
        for year in range(1, 4):
            total = round(
                sems.get((year, "fall"), {}).get("total", 0.0)
                + sems.get((year, "spring"), {}).get("total", 0.0),
                4,
            )
            target = cfg.get("target_annual_load", 4.0)
            hard_cap = cfg.get("junior_faculty_hard_cap", 2.0) * 2
            soft_cap = cfg.get("senior_faculty_soft_cap", 2.0) * 2
            cap = hard_cap if (f and f.is_junior()) else soft_cap
            if total > cap + 1.0:
                status = "red"
            elif total > cap:
                status = "yellow-over"
            elif total >= target - 0.5:
                status = "green"
            elif total > 0:
                status = "yellow-under"
            else:
                status = "empty"
            annual_loads[fname][year] = {"total": total, "status": status}

    violations = compute_violations(faculty_list, loads, plan, courses, cfg)

    return render_template(
        "planner.html",
        semesters=semesters,
        faculty_list=faculty_list,
        loads_by_sem=loads_by_sem,
        annual_loads=annual_loads,
        violations=violations,
        cfg=cfg,
    )


@app.route("/assign", methods=["POST"])
def assign():
    data = request.get_json()
    slot_id = data.get("slot_id", "")
    faculty_name = data.get("faculty_name", "")

    parts = slot_id.split("__")
    if len(parts) != 4:
        return jsonify({"error": "Invalid slot_id"}), 400

    course_code, year, season, section = parts[0], int(parts[1]), parts[2], int(parts[3])

    plan = load_plan()
    existing = plan.get_assignment(course_code, year, season, section)
    if existing and existing.locked:
        return jsonify({"error": "Assignment is locked"}), 403

    faculty = next((f for f in faculty_list if f.name == faculty_name), None)
    if faculty is None:
        return jsonify({"error": "Unknown faculty"}), 400

    course = courses.get(course_code)
    if course is None:
        return jsonify({"error": "Unknown course"}), 400
    if not course.is_placeholder and not faculty.can_teach_course(course_code):
        return jsonify({"error": "Faculty not qualified for this course"}), 400

    flavor = None
    if course_code == "sci10":
        for fl in ("health", "neuro", "earth"):
            if faculty.can_teach.get(f"sci10_{fl}"):
                flavor = fl
                break

    plan.set_assignment(Assignment(
        faculty_name=faculty_name,
        course_code=course_code,
        year=year,
        semester=season,
        section_number=section,
        locked=False,
        manual=True,
        flavor=flavor,
    ))
    save_plan(plan)
    return jsonify({"ok": True})


@app.route("/unassign", methods=["POST"])
def unassign():
    data = request.get_json()
    slot_id = data.get("slot_id", "")

    parts = slot_id.split("__")
    if len(parts) != 4:
        return jsonify({"error": "Invalid slot_id"}), 400

    course_code, year, season, section = parts[0], int(parts[1]), parts[2], int(parts[3])

    plan = load_plan()
    existing = plan.get_assignment(course_code, year, season, section)
    if existing and existing.locked:
        return jsonify({"error": "Assignment is locked"}), 403

    plan.remove_assignment(course_code, year, season, section)
    save_plan(plan)
    return jsonify({"ok": True})


def build_diagnostics(plan: Plan) -> dict:
    loads = all_faculty_loads(faculty_list, plan.assignments, courses, cfg)
    semesters = build_grid(plan)

    # 1. Coverage per semester
    coverage = []
    for sem in semesters:
        total = sem["total_sections"]
        filled = sum(
            1 for g in sem["groups"]
            for s in g["sections"]
            if s["assignment"] is not None
        )
        coverage.append({
            "label": sem["label"],
            "filled": filled,
            "total": total,
            "pct": round(filled / total * 100) if total else 0,
        })

    # 2. Unfilled sections
    unfilled = []
    for sem in semesters:
        for g in sem["groups"]:
            for s in g["sections"]:
                if s["assignment"] is None:
                    unfilled.append({
                        "semester_label": sem["label"],
                        "course_name": g["course"].display_name,
                        "section": s["section_number"],
                        "is_placeholder": g["course"].is_placeholder,
                    })

    # 3. Faculty load table — sorted juniors first then seniors, alphabetical within
    target = cfg.get("target_annual_load", 4.0)
    hard_cap = cfg.get("junior_faculty_hard_cap", 2.0) * 2
    soft_cap = cfg.get("senior_faculty_soft_cap", 2.0) * 2

    faculty_loads_table = []
    for f in sorted(faculty_list, key=lambda x: (x.rank, x.name)):
        row = {"name": f.name, "rank": "J" if f.is_junior() else "S", "area": f.area, "years": {}}
        cap = hard_cap if f.is_junior() else soft_cap
        for year in range(1, 4):
            fall_t = loads.get(f.name, {}).get((year, "fall"), {}).get("total", 0.0)
            spring_t = loads.get(f.name, {}).get((year, "spring"), {}).get("total", 0.0)
            annual = round(fall_t + spring_t, 2)
            if annual > cap + 1.0:
                status = "red"
            elif annual > cap:
                status = "yellow-over"
            elif annual >= target - 0.5:
                status = "green"
            elif annual > 0:
                status = "yellow-under"
            else:
                status = "empty"
            row["years"][year] = {
                "fall": round(fall_t, 2),
                "spring": round(spring_t, 2),
                "annual": annual,
                "status": status,
            }
        faculty_loads_table.append(row)

    # 4. Bottleneck courses (non-placeholder only), sorted by sections/faculty ratio desc
    bottlenecks = []
    for code, course in courses.items():
        if course.is_placeholder:
            continue
        qualified = sum(1 for f in faculty_list if f.can_teach_course(code))
        if code == "sci10":
            total_sections = sum(
                _sci10_count(plan, y, s)
                for y in range(plan.year_range[0], plan.year_range[1] + 1)
                for s in ("fall", "spring")
            )
        else:
            total_sections = (
                course.sections_per_semester.get("fall", 0)
                + course.sections_per_semester.get("spring", 0)
            ) * 3
        ratio = round(total_sections / qualified, 2) if qualified else None
        bottlenecks.append({
            "name": course.display_name,
            "qualified": qualified,
            "total_sections": total_sections,
            "ratio": ratio,
        })
    bottlenecks.sort(key=lambda x: (x["ratio"] is None, -(x["ratio"] or 0)))

    # 5. Junior new-prep counts per year (brand-new = zero prior history, matching constraint logic)
    junior_preps = []
    for f in faculty_list:
        if not f.is_junior():
            continue
        cumulative = dict(f.prior_teaching_counts)
        years_data = {}
        for year in range(1, 4):
            year_new = set()
            for sem in ("fall", "spring"):
                sem_assigns = [
                    a for a in plan.assignments
                    if a.faculty_name == f.name and a.year == year and a.semester == sem
                ]
                for a in sem_assigns:
                    ck = count_key(a)
                    if cumulative.get(ck, 0) == 0:
                        year_new.add(a.course_code)
                    cumulative[ck] = cumulative.get(ck, 0) + 1
            years_data[year] = {
                "count": len(year_new),
                "names": [courses[c].display_name for c in year_new if c in courses],
            }
        junior_preps.append({"name": f.name, "years": years_data})

    return {
        "coverage": coverage,
        "unfilled": unfilled,
        "faculty_loads_table": faculty_loads_table,
        "bottlenecks": bottlenecks,
        "junior_preps": junior_preps,
    }


@app.route("/diagnostics")
def diagnostics():
    plan = load_plan()
    data = build_diagnostics(plan)
    return render_template("diagnostics.html", **data)


@app.route("/solve", methods=["POST"])
def solve_route():
    plan = load_plan()
    new_assignments = run_solver(faculty_list, courses, plan, cfg, plan.year_range)
    # Drop previous solver assignments (not locked, not manual), keep locked + manual
    plan.assignments = [a for a in plan.assignments if a.locked or a.manual]
    for a in new_assignments:
        plan.assignments.append(a)
    save_plan(plan)
    return redirect(url_for("index"))


@app.route("/clear_solver", methods=["POST"])
def clear_solver():
    plan = load_plan()
    plan.assignments = [a for a in plan.assignments if a.locked or a.manual]
    save_plan(plan)
    return redirect(url_for("index"))


@app.route("/lock", methods=["POST"])
def lock():
    data = request.get_json()
    slot_id = data.get("slot_id", "")
    parts = slot_id.split("__")
    if len(parts) != 4:
        return jsonify({"error": "Invalid slot_id"}), 400
    course_code, year, season, section = parts[0], int(parts[1]), parts[2], int(parts[3])
    plan = load_plan()
    existing = plan.get_assignment(course_code, year, season, section)
    if existing is None:
        return jsonify({"error": "No assignment to lock"}), 404
    existing.locked = True
    save_plan(plan)
    return jsonify({"ok": True})


@app.route("/unlock", methods=["POST"])
def unlock():
    data = request.get_json()
    slot_id = data.get("slot_id", "")
    parts = slot_id.split("__")
    if len(parts) != 4:
        return jsonify({"error": "Invalid slot_id"}), 400
    course_code, year, season, section = parts[0], int(parts[1]), parts[2], int(parts[3])
    plan = load_plan()
    existing = plan.get_assignment(course_code, year, season, section)
    if existing is None:
        return jsonify({"error": "No assignment to unlock"}), 404
    existing.locked = False
    save_plan(plan)
    return jsonify({"ok": True})


@app.route("/lock_all", methods=["POST"])
def lock_all():
    plan = load_plan()
    for a in plan.assignments:
        a.locked = True
    save_plan(plan)
    return redirect(url_for("index"))


@app.route("/unlock_all", methods=["POST"])
def unlock_all():
    plan = load_plan()
    for a in plan.assignments:
        a.locked = False
    save_plan(plan)
    return redirect(url_for("index"))


@app.route("/set_flavor", methods=["POST"])
def set_flavor():
    data = request.get_json()
    slot_id = data.get("slot_id", "")
    flavor = data.get("flavor", "")

    if flavor not in ("health", "neuro", "earth"):
        return jsonify({"error": "Invalid flavor"}), 400

    parts = slot_id.split("__")
    if len(parts) != 4:
        return jsonify({"error": "Invalid slot_id"}), 400

    course_code, year, season, section = parts[0], int(parts[1]), parts[2], int(parts[3])
    if course_code != "sci10":
        return jsonify({"error": "Only sci10 assignments have flavors"}), 400

    plan = load_plan()
    existing = plan.get_assignment(course_code, year, season, section)
    if existing is None:
        return jsonify({"error": "No assignment found"}), 404

    faculty = next((f for f in faculty_list if f.name == existing.faculty_name), None)
    if faculty and not faculty.can_teach.get(f"sci10_{flavor}", False):
        return jsonify({"error": f"{existing.faculty_name} is not qualified for sci10 {flavor}"}), 400

    existing.flavor = flavor
    save_plan(plan)
    return jsonify({"ok": True})


@app.route("/set_sci10_sections", methods=["POST"])
def set_sci10_sections():
    data = request.get_json()
    year = data.get("year")
    season = data.get("season")
    delta = data.get("delta")

    if not isinstance(year, int) or season not in ("fall", "spring"):
        return jsonify({"error": "Invalid year or season"}), 400
    if delta not in (1, -1):
        return jsonify({"error": "delta must be 1 or -1"}), 400

    plan = load_plan()
    current = _sci10_count(plan, year, season)
    new_count = current + delta

    if new_count < 1 or new_count > 20:
        return jsonify({"error": "Section count must be between 1 and 20"}), 400

    if delta == -1 and plan.get_assignment("sci10", year, season, current) is not None:
        return jsonify({"error": "Cannot remove a section that has an assignment"}), 400

    plan.sci10_section_overrides[f"{year}__{season}"] = new_count
    save_plan(plan)
    return jsonify({"ok": True, "count": new_count})


@app.route("/update_config", methods=["POST"])
def update_config():
    global cfg
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data"}), 400

    scalar_keys = [
        "junior_faculty_hard_cap", "senior_faculty_soft_cap",
        "junior_new_lab_preps_per_year_max", "new_prep_bonus_count",
        "new_prep_weight", "foundational_experienced_weight",
        "extra_section_weight_multiplier", "target_annual_load",
    ]
    int_keys = {"junior_new_lab_preps_per_year_max", "new_prep_bonus_count"}
    weight_keys = [
        "section_coverage", "junior_new_preps", "load_balance",
        "senior_over_cap", "sci10_flavor_diversity", "senior_takes_new_preps",
    ]

    new_cfg = dict(cfg)
    for k in scalar_keys:
        if k in data:
            try:
                new_cfg[k] = int(data[k]) if k in int_keys else float(data[k])
            except (ValueError, TypeError):
                return jsonify({"error": f"Invalid value for {k}"}), 400

    new_weights = dict(cfg.get("objective_weights", {}))
    for k in weight_keys:
        if k in data.get("objective_weights", {}):
            try:
                new_weights[k] = int(data["objective_weights"][k])
            except (ValueError, TypeError):
                return jsonify({"error": f"Invalid value for objective weight {k}"}), 400
    new_cfg["objective_weights"] = new_weights

    cfg = new_cfg
    with open(CONFIG_FILE, "w") as f:
        yaml.dump(cfg, f, default_flow_style=False, sort_keys=False)

    return jsonify({"ok": True})


@app.route("/export")
def export():
    plan = load_plan()
    loads = all_faculty_loads(faculty_list, plan.assignments, courses, cfg)
    diag = build_diagnostics(plan)
    faculty_by_name = {f.name: f for f in faculty_list}

    wb = openpyxl.Workbook()

    # ── Sheet 1: Plan ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Plan"

    plan_headers = [
        "Faculty", "Rank", "Course", "Year", "Semester", "Section",
        "Flavor", "Locked", "Manual", "Base Weight", "Actual Weight",
    ]
    for col, h in enumerate(plan_headers, 1):
        ws.cell(1, col, h).font = _BOLD

    # Build weight lookup: (faculty, course, year, sem, section) → item dict
    weight_lookup = {}
    for fname, sems in loads.items():
        for (year, sem), data in sems.items():
            for item in data.get("items", []):
                key = (fname, item["course_code"], year, sem, item["section_number"])
                weight_lookup[key] = item

    sorted_assignments = sorted(
        plan.assignments,
        key=lambda a: (a.year, 0 if a.semester == "fall" else 1, a.course_code, a.section_number),
    )
    for r, a in enumerate(sorted_assignments, 2):
        f = faculty_by_name.get(a.faculty_name)
        course = courses.get(a.course_code)
        wt = weight_lookup.get((a.faculty_name, a.course_code, a.year, a.semester, a.section_number), {})
        row = [
            a.faculty_name,
            "J" if (f and f.is_junior()) else "S",
            course.display_name if course else a.course_code,
            a.year,
            a.semester.capitalize(),
            a.section_number,
            a.flavor or "",
            "Yes" if a.locked else "No",
            "Yes" if a.manual else "No",
            wt.get("base_weight", ""),
            wt.get("actual_weight", ""),
        ]
        for col, val in enumerate(row, 1):
            ws.cell(r, col, val)

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    # ── Sheet 2: Gap Report ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Gap Report")
    row = 1

    def gap_header(label):
        nonlocal row
        ws2.cell(row, 1, label).font = _BOLD
        row += 1

    def gap_row(values, fills=None):
        nonlocal row
        for col, val in enumerate(values, 1):
            cell = ws2.cell(row, col, val)
            if fills and col <= len(fills) and fills[col - 1]:
                cell.fill = fills[col - 1]
        row += 1

    def gap_subheader(cols):
        nonlocal row
        for col, h in enumerate(cols, 1):
            ws2.cell(row, col, h).font = _BOLD
        row += 1

    # Coverage
    gap_header("Coverage")
    gap_subheader(["Semester", "Filled", "Total", "%"])
    for c in diag["coverage"]:
        gap_row([c["label"], c["filled"], c["total"], c["pct"]])
    row += 1

    # Unfilled sections
    gap_header("Unfilled Sections")
    gap_subheader(["Semester", "Course", "Section", "Placeholder"])
    for u in diag["unfilled"]:
        gap_row([
            u["semester_label"], u["course_name"], u["section"],
            "Yes" if u["is_placeholder"] else "No",
        ])
    row += 1

    # Faculty loads
    gap_header("Faculty Loads")
    load_headers = ["Faculty", "Rank"]
    for yr in [1, 2, 3]:
        load_headers += [f"Y{yr} Fall", f"Y{yr} Spring", f"Y{yr} Annual"]
    gap_subheader(load_headers)
    for frow in diag["faculty_loads_table"]:
        values = [frow["name"], frow["rank"]]
        fills = [None, None]
        for yr in [1, 2, 3]:
            yd = frow["years"].get(yr, {})
            fill = _STATUS_FILL.get(yd.get("status", "empty"))
            values += [yd.get("fall", 0), yd.get("spring", 0), yd.get("annual", 0)]
            fills += [fill, fill, fill]
        gap_row(values, fills=fills)
    row += 1

    # Bottleneck courses
    gap_header("Bottleneck Courses")
    gap_subheader(["Course", "Qualified Faculty", "Total Sections (3yr)", "Sections/Faculty"])
    for b in diag["bottlenecks"]:
        gap_row([b["name"], b["qualified"], b["total_sections"], b["ratio"]])
    row += 1

    # Junior new-prep counts
    gap_header("Junior New-Prep Counts")
    gap_subheader(["Faculty", "Y1 Count", "Y1 Courses", "Y2 Count", "Y2 Courses", "Y3 Count", "Y3 Courses"])
    for jp in diag["junior_preps"]:
        values = [jp["name"]]
        for yr in [1, 2, 3]:
            yd = jp["years"].get(yr, {})
            values += [yd.get("count", 0), ", ".join(yd.get("names", []))]
        gap_row(values)

    for col in ws2.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        download_name="cmc_plan.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/faculty/<name>")
def faculty_detail(name):
    faculty = next((f for f in faculty_list if f.name == name), None)
    if faculty is None:
        return "Faculty not found", 404
    return render_template("faculty_detail.html", faculty=faculty)


if __name__ == "__main__":
    app.run(debug=True, port=5000)
